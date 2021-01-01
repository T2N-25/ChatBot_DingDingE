#모듈 및 기초설정
from konlpy.tag import Komoran
okt = Komoran()
import openpyxl as opx
wb = opx.load_workbook('resource/datafile2.xlsx') # wb = workbook
ws = wb['Sheet1'] # ws = worksheet
import time as t
import json
from datetime import datetime
from googletrans import Translator
from bs4 import BeautifulSoup as bs
from pprint import pprint
import requests
import pyowm
import random
#==
RSP = ["바위","가위","보"]
#함수 모음

def 베스킨라빈스():


    def 메롱():
        global win
        global lose
        global turn

        n = random.randint(31,50)
        k = random.randint(3,5)

        r = (n-1) % (k+1) #처음 시작할때 필승 숫자
        i = 1


        print("""*규칙
        당신과 딩딩이가 번갈아가며 1부터 순서대로 숫자를 부릅니다.
        {}를 부르는 사람이 지고, 한 번에 {}개의 숫자를 부를 수 있습니다.
        채팅창에 n을 적으면 n개의 숫자가 나열됩니다.
        ex) 2를 적으면 1 2, 그 다음 3을 적으면 3 4 5
        먼저 시작하려면 1, 딩딩이부터 시작하려면 2, 그만하려면 3을 입력해주세요.
        """.format(n, k))
        while True:
            try:
                turn = int(input())
                break
            except:
                print("다시 입력하세요")

        if turn == 3:
            return

        while True:
            if turn == 1:

            # 당신차례
                while True:
                    try:
                        count = int(input('>>> ')) #당신이 뽑을 숫자 개수 입력
                        print() 
                        break
                    except:
                        print("다시 입력하세요")

                if 1 <= count <= k:
                    print("사용자: ", end="")
                    for m in range(count):
                        print(i, end=" ")
                        if i == n:
                            print("당신이 졌습니다ㅠㅠ")
                            lose += 1
                            return
                        i+=1



                    print()
                    turn = 2

                else:
                    print("1부터", k, "까지의 수를 입력하세요.")



            else:    # 딩딩이 턴
                if (i-1)%(k+1) == r:
                    count = random.randint(1, k)

                else:
                    count = (r-(i-1)%(k+1))%(k+1)

                print("딩딩이: ", end="")

                for m in range(count): # count개를 print
                    print(i, end=" ")
                    if i == n:
                        print("당신이 이겼습니다!!")
                        win += 1
                        return
                    i+=1

                print()

                turn = 1

    while True:
        메롱()
        if win == 0:
            win_rate = 0
        else:
            win_rate = win / (win + lose) * 100
        print("지금까지 당신의 승률은 {}% 입니다!\n당신의 승리 : {}\n딩딩이 승리 : {}".format(win_rate,win,lose))
        print()

        if turn == 3:
            break
            
    print("게임을 종료합니다.")


def 가위바위보():
    while True:
        player = input("가위, 바위, 보 중 어느것을 내시겠어요?\n>>> ")
        computer = random.choice(RSP)
        if player == computer:
            print("당신 : {} \n딩딩이 : {}".format(player,computer))
            print("비겼어요!")
            break
        elif player == "바위":
            if computer == "가위":
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("바위를 냈군요! 딩딩이가 가위를 냈으니 딩딩이가 졌어요~~")
                break
            else:
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("바위를 냈군요! 딩딩이가 보를 냈으니 당신이 졌어요~~")
                break
        elif player == "보":
            if computer == "바위":
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("보를 냈군요! 딩딩이가 바위를 냈으니 딩딩이가 졌어요~~")
                break
            else:
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("보를 냈군요! 딩딩이가 가위를 냈으니 당신이 졌어요~~")
                break
        elif player == "가위":
            if computer == "보":
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("가위를 냈군요! 딩딩이가 보를 냈으니 딩딩이가 졌어요~~")
                break
            else:
                print("당신 : {} \n딩딩이 : {}".format(player,computer))
                print("가위를 냈군요! 딩딩이가 바위를 냈으니 당신이 졌어요~~")
                break
        else:
            print("잘못냈어요!")
        print()
def Save(List, c, cn):
    ws[cn] = len(List)+3
    print(len(List)+3)
    print(List)
    for i in range(3,len(List)+3):
        S = c + str(i)
        n = i - 3
        ws[S] = List[n]

def SkillWord(word):
    if word in Skill:
        return True
    else:
        return False
def BanWord(word):
    if word in Qx:
        return True
    else:
        return False
def FillWord(word):
    if word in Fa:
        return True
    else:
        return False
def BasicWord(word):
    if word in Qs:
        return True
    else:
        return False
def Basic2Word(word):
    if word in Q2s:
        return True
    else:
        return False
#==
def Print_City_Temp(City_ID):
        obs = owm.weather_at_id(City_ID)  
        
        L = obs.get_location()
        City_name = L.get_name()
        
        W = obs.get_weather()
        Status = W.get_status()
        
        Temp = W.get_temperature(unit='celsius')
        print('웅상의 현재 날씨는 '+ Status + '이고 온도는 ' + str(Temp['temp']) + ' 입니다!')
#==

#==
def format_number(number):
    return '{:,d}'.format(number)


def get_country_data(country):
    country_data_url = "https://corona.lmao.ninja/v3/covid-19/countries/" + country
    res = requests.get(country_data_url).text
    if 'Country not found or doesn' in res:
        print("나라이름을 잘못 작성하셨습니다.")
        print(country, " 는 존재하지 않는 나라이름이거나, 영문으로 작성되지 않았습니다.")
    else:
        country_corona_info = json.loads(res)
        print("[ %s Corona Data ]" % country)
        print("추가 확진자: +%s" % format_number(country_corona_info["todayCases"]))
        print("추가 사망자: +%s" % format_number(country_corona_info["todayDeaths"]))
        print("확진자: %s" % format_number(country_corona_info["cases"]))
        print("사망자: %s" % format_number(country_corona_info["deaths"]))
        print("격리해제: %s" % format_number(country_corona_info["recovered"]))
        print("격리중: %s" % format_number(country_corona_info["active"]))
        return country_corona_info


def get_global_data():
    global_data_url = "https://corona.lmao.ninja/v3/covid-19/all"
    res = requests.get(global_data_url).text
    country_corona_info = json.loads(res)
    print("[ Global Corona Data ]")
    print("확진자: %s" % format_number(country_corona_info["cases"]))
    print("사망자: %s" % format_number(country_corona_info["deaths"]))
    print("격리해제: %s" % format_number(country_corona_info["recovered"]))
    print("(%s 기준)" % datetime.fromtimestamp(country_corona_info["updated"]/1000.0))
    return country_corona_info


def main():
    stringCountry = 'S. korea'
    get_country_data(stringCountry)
    print()
#==
#==


Qx = []
Fs = []
Fa = []
Qs = []
As = []
Hs = []
Q2s = [] # Q S
H2s = [] # R T
# 리스트에 값을 넣어두고, 프로그램을 종료하거나 시작할때 값을 리스트에 읽고 쓰기 동기형식!

#기본 설정
API_Key = '52a192d87186e871d8ad7c1300c3730d'
owm = pyowm.OWM(API_Key)
Fill = ['기쁨','분노','슬픔','즐거움','사랑','증오','욕망']
Skill = ['계산','번역','미세먼지','코로나','날씨','날짜','시간','가위바위보','베스킨라빈스']
#==

#값 읽기

G = ws['G2']
for i in range(3,G.value): # Qx
    C = "C" + str(i)
    Qx.append(ws[C].value)
#print(Qx)

M = ws['M2']
for i in range(3,M.value): # Fs
    K = "K" + str(i)
    Fs.append(ws[K].value)
#print(Fs)
N = ws['N2']
for i in range(3,N.value): # Fa
    L = "L" + str(i)
    Fa.append(ws[L].value)
#print(Fa)

I = ws['I2']
for i in range(3,I.value): # Qs
    E = "E" + str(i)
    Qs.append(ws[E].value)
#print(Qs)

J = ws['J2']
for i in range(3,J.value): # As
    F = "F" + str(i)
    As.append(ws[F].value)
#print(As)

P = ws['P2']
for i in range(3,P.value): # As
    O = "O" + str(i)
    Hs.append(ws[O].value)
#print(Hs)

S = ws['S2']
for i in range(3,S.value): # Q2s
    Q = "Q" + str(i)
    Q2s.append(ws[Q].value)
#print(Q2s)

T = ws['T2']
for i in range(3,T.value): # H2s
    R = "R" + str(i)
    H2s.append(ws[R].value)
#print(H2s)
#==
while True:
    an = input('[실행/가르치기/저장/종료]\n>>> ')
    if an == '가르치기':
        pss = input('딩코부원만 가르칠 수 있습니다.\n패스워드를 입력해 주세요.\n>>> ')
        if pss == "DcCbPassWord":
            while True:
                wp = input('[금칙어필터/기분필터/일상대화/그만가르치기] 중 어떤 기능을 가르치겠습니까?\n>>> ')

                
                if wp == '금칙어필터':
                    IsSet = False
                    Qxa = input('금칙어를 입력해주세요!\n>>> ')
                    for i in Qx:
                        if i == Qxa:
                            IsSet = True
                    if IsSet:
                        print(Qxa,'는 이미 딩딩이가 알고있는 금칙어 입니다!')
                    else:
                        Qx.append(Qxa)
                        print(Qxa,'를 금칙어로 가르쳤습니다.\n봇 딩딩이는 실행기능중, 해당 단어가 포함된 말을 들으면 금칙어필터를 가동할것입니다.')

                        
                elif wp == '기분필터':
                    IsSet = False
                    Fxa = input('어떤말을 들었나요?\n>>> ')
                    for i in Fa:
                        if i == Fxa:
                            IsSet = True
                    if IsSet:
                        print(Fxa,'는 이미 딩딩이가 알고있는 기분대화 입니다!')
                    else:
                        while True:
                            Fsa = input("이 말을 하는 사람의 기분은 어떤가요? ['기쁨','분노','슬픔','즐거움','사랑','증오','욕망']\n>>> ")
                            if Fsa in Fill:
                                Fs.append(Fsa)
                                Fa.append(Fxa)
                                print(Fxa,' 기분대화를',Fsa,'기분으로 가르쳤습니다.\n봇 딩딩이는 실행기능중, 해당 단어가 포함된 말을 들으면 기분필터를 가동할것입니다.')
                                break
                            else:
                                print(Fsa,'라는 기분은 등록되어 있지 않습니다!')


                        
                elif wp == '일상대화':
                    IsSet = False
                    Qsa = input('어떤명사를 들었나요?\n>>> ') #질문1
                    Q2sa = input('어떤 형용사나 동사를 들었나요?\n>>> ') #질문2
                    for i in Qs:
                        if i == Qsa:
                            IsSet = True
                    for i in Q2s:
                        if i == Q2sa:
                            Is2Set = True
                    Asa = input('그말을 들은 딩딩이는 어떻게 대답해야 할까요?\n>>> ') #답 Hs 형태소
                    if IsSet == True and Is2Set == True:
                        dans = 0
                        for dsz in Qs:
                            if Qsa == dsz:
                                dnqs = dans
                            dans += 1
                        As[dnqs] = As[dnqs]+'§'+Asa
                        print(Qsa,'와,',Q2sa,'을/를 들으면',Asa,'라고 대답한다!\n봇 딩딩이는 일상대화중, 해당 단어가 포함된 말을 들으면 해당 하는 답을 제시할것입니다.')
                    else:
                        Hsa = input('명사의 형태소는 무엇인가요?\n>>> ')
                        H2sa = input('형용사나 동사의 형태소는 무엇인가요?\n>>> ')
                        Qs.append(Qsa)
                        As.append(Asa)
                        Hs.append(Hsa)
                        Q2s.append(Q2sa)
                        H2s.append(H2sa)
                        print(Qsa,'와,',Q2sa,'을/를 들으면',Asa,'라고 대답한다!(형태소는',Hsa,',',H2sa,')\n봇 딩딩이는 일상대화중, 해당 단어가 포함된 말을 들으면 해당 하는 답을 제시할것입니다.')
                        
                elif wp == '그만가르치기':
                    break
                else:
                    print('[금칙어필터/일상대화/그만가르치기] 중에서 입력해주세요.')
        else:
            print("패스워드가 틀렸습니다.")
    elif an == '실행': # '종료' -> 기능대화 -> 금칙어필터 -> 기분필터 -> 일상필터 -> 모르는것
        print('3초뒤 딩딩이를 불러냅니다!')
        t.sleep(3)
        print('안녕하세요! 저는 딩딩이에요!')
        while True:
            dda = input('>>> ')
            if dda == '종료':
                break
            ddan = okt.morphs(dda)
            Lddan = list(reversed(ddan))
            ddan2 = okt.pos(dda)
            RLSMD = 0
            Lddan2 = list(reversed(ddan2))
            for Ri in Lddan:
                Ans = Ri
                if SkillWord(Ri):
#================================================
                    if Ans == "날씨": # 날씨
#================================================
                        Print_City_Temp(1912205)
#================================================
                    elif Ans == "날짜" or Ans == "시간": # 날짜와 시간
#================================================
                        now = datetime.now()
                        print ("%s년 %s월 %s일 %s시 %s분 %s초입니다!" %(now.year, now.month, now.day, now.hour, now.minute, now.second))
#================================================
                    elif Ans == "미세먼지": # 미세먼지와 초미세먼지
#================================================
                        html = requests.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query=%EB%AF%B8%EC%84%B8%EB%A8%BC%EC%A7%80&oquery=%EB%AF%B8%EC%84%B8%EB%A8%BC%EC%A7%80+%EC%B4%88%EB%AF%B8%EC%84%B8%EB%A8%BC%EC%A7%80&tqi=U9mb1dprvN8ssvgrsThssssss3C-095150')
                        # pprint(html.text)
                        soup = bs(html.text,'html.parser')
                        data1 = soup.find('div',{'class':'tb_scroll'})
                        #pprint(data1)
                        data2 = data1.findAll('tbody')
                        # pprint(data2)
                        fine_dust = data2[0].findAll('span')
                        print("대한민국 경상남도 양산시의 미세먼지 지수는",fine_dust[27].text,"입니다!")#,"초미세먼지 지수는",ultra_fine_dust,"입니다!")
#================================================
                    elif Ans == "번역": # 번역
#================================================
                        TBTB = dda.split('"')
                        DGFG = okt.pos(TBTB[1])
                        GTDG = 0
                        for i in DGFG:
                            if 'SL' in i:
                                GTDG = 1
                        if GTDG == 1:
                            source = 'en'
                            destion = 'ko'
                        else:
                            source = 'ko'
                            destion = 'en'
                        forTranslatorString = TBTB[1]
    
                        translator = Translator()
                        value = translator.translate(forTranslatorString, src=source, dest=destion)
                     
                        #print(value)
                        #print(value.src)  # 변환할 언어
                        #print(value.dest)  # 변환될 언어
                        print(TBTB[1], '를 번역했어요!')  # 변환 결과
                        print('=>' , value.text)
#================================================
                    elif Ans == "계산": # 계산기
#================================================
                        s = input("값을 입력해 주세요 : ")
                        print("결과 : {}".format(eval(s)))
#================================================
                    elif Ans == "코로나": # 코로나 확진자
#================================================
                        print('코로나 확진자 시스템을 시작합니다!')
                        t.sleep(1)
                        get_global_data()
                        print()
                        if __name__ == '__main__':
                            main()
#================================================
                    elif Ans == "가위바위보": # 가위바위보
#================================================
                        가위바위보()
#================================================
                    elif "베스킨라빈스" in dda: # 베스킨라빈스
#================================================
                        win = 0
                        lose = 0
                        베스킨라빈스()
                    RLSMD = 1
                    break
            LOPIO = []
            for EDEN in Lddan2:
                LOPIO.append(EDEN[1])
            #==============================================
            NNP = []
            NNG = []
            VV = []
            NP = []
            VA = []
            NA = []
            
            n = 0
            for i in LOPIO:
                if i == 'NNP':
                    NNP.append(n)
                if i == 'NNG':
                    NNG.append(n)
                if i == 'VV':
                    VV.append(n)
                if i == 'NP':
                    NP.append(n)
                if i == 'VA':
                    VA.append(n)
                if i == 'NA':
                    NA.append(n)
                n += 1
            HTS = NNP + NNG
            H2TS = VV + NP + VA
            for Zi in HTS:
                if RLSMD == 1:
                    break
                Ri = Lddan[Zi]
                if BanWord(Ri):
                    print('그런말 하면 안되죠!')
                    break
                elif BasicWord(Ri):
                    dand = 0
                    for dz in Qs:
                        if Ri == dz:
                            dnqd = dand
                        dand += 1
                    #print('dnqd :',dnqd)
#============================================
                    TRST = 0
                    for Di in H2TS:
                        Gi = Lddan[Di]
                        #print('Gi :', Gi)
                        if Basic2Word(Gi):
                            TRST = 1
                            if '§' in As[dnqd]:
                                GED = As[dnqd].split('§')
                                print(random.choice(GED))
                            else:
                                print(As[dnqd])
                            break
                    if TRST == 0:
                        if '§' in As[dnqd]:
                            GED = As[dnqd].split('§')
                            print(random.choice(GED))
                        else:
                            print(As[dnqd])
            for Yi in NA:
                if 'ㅋㅋ' in Lddan[Yi]:
                    print('키득키득')
                elif 'ㅠㅠ' in Lddan[Yi]:
                    print('우어ㅠㅠ')
                elif 'ㅋㅎ' in Lddan[Yi]:
                    print('ㅋㅎㅋㅎ')
                elif 'ㅎㅎ' in Lddan[Yi]:
                    print('ㅎㅎㅎㅎ')
            
    elif an == "저장":
        print("저장을 시작합니다!")

        
        Save(Qx,'C','G2')
        Save(Fs,'K','M2')
        Save(Fa,'L','N2')
        Save(Qs,'E','I2')
        Save(As,'F','J2')
        Save(Hs,'O','P2')
        Save(Q2s,'Q','S2')
        Save(H2s,'R','T2')
        wb.save('resource/datafile.xlsx')


        print("저장이 완료되었습니다.")
    elif an == "종료":
        print("저장후 프로그램이 종료됩니다.")
        t.sleep(1)
        print("저장을 시작합니다!")

        
        Save(Qx,'C','G2')
        Save(Fs,'K','M2')
        Save(Fa,'L','N2')
        Save(Qs,'E','I2')
        Save(As,'F','J2')
        Save(Hs,'O','P2')
        Save(Q2s,'Q','S2')
        Save(H2s,'R','T2')
        wb.save('resource/datafile.xlsx')

        
        print("저장이 완료되었습니다.")
        t.sleep(1)
        print("5초후 프로그램이 종료됩니다.")
        t.sleep(5)
        break
    else:
        print('[실행/가르치기/저장/종료] 중에서만 입력해주세요!')
